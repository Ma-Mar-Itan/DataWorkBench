"""Workbook reader for loading Excel and CSV files."""

import pandas as pd
from typing import Dict, Any, Optional, Tuple
from pathlib import Path


def load_workbook(file_path: str) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    """
    Load a workbook from file.
    
    Args:
        file_path: Path to the .xlsx or .csv file
        
    Returns:
        Tuple of (dict of sheet_name -> DataFrame, metadata dict)
    """
    path = Path(file_path)
    extension = path.suffix.lower()
    
    if extension == ".xlsx":
        return _load_excel(path)
    elif extension == ".csv":
        return _load_csv(path)
    else:
        raise ValueError(f"Unsupported file type: {extension}")


def _load_excel(path: Path) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    """Load an Excel file with all sheets."""
    import openpyxl
    
    # Get sheet names first
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    
    # Load each sheet into a DataFrame
    sheets = {}
    rows_per_sheet = {}
    columns_per_sheet = {}
    total_cells = 0
    
    for sheet_name in sheet_names:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
        sheets[sheet_name] = df
        rows_per_sheet[sheet_name] = len(df)
        columns_per_sheet[sheet_name] = len(df.columns)
        total_cells += len(df) * len(df.columns)
    
    wb.close()
    
    metadata = {
        "file_name": path.name,
        "num_sheets": len(sheet_names),
        "sheet_names": sheet_names,
        "rows_per_sheet": rows_per_sheet,
        "columns_per_sheet": columns_per_sheet,
        "total_cells": total_cells,
    }
    
    return sheets, metadata


def _load_csv(path: Path) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    """Load a CSV file as a single-sheet workbook."""
    df = pd.read_csv(path, dtype=str)
    
    metadata = {
        "file_name": path.name,
        "num_sheets": 1,
        "sheet_names": ["Sheet1"],
        "rows_per_sheet": {"Sheet1": len(df)},
        "columns_per_sheet": {"Sheet1": len(df.columns)},
        "total_cells": len(df) * len(df.columns),
    }
    
    return {"Sheet1": df}, metadata


def save_workbook(sheets: Dict[str, pd.DataFrame], output_path: str, 
                  original_path: Optional[str] = None) -> None:
    """
    Save cleaned data back to an Excel file.
    
    Args:
        sheets: Dict of sheet_name -> DataFrame
        output_path: Path for the output file
        original_path: Optional path to original file (for preserving format)
    """
    output = Path(output_path)
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            # Ensure sheet name is valid (Excel has 31 char limit)
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    
    # If we have an original file, try to preserve some formatting
    if original_path and Path(original_path).exists():
        _preserve_formatting(original_path, output)


def _preserve_formatting(original_path: str, output_path: str) -> None:
    """
    Attempt to preserve basic formatting from original workbook.
    
    This is a simplified implementation that preserves:
    - Sheet order
    - Basic structure
    """
    import openpyxl
    
    try:
        # Load original for reference
        original_wb = openpyxl.load_workbook(original_path)
        output_wb = openpyxl.load_workbook(output_path)
        
        # Remove default sheet if it exists
        if "Sheet" in output_wb.sheetnames and len(output_wb.sheetnames) > 1:
            del output_wb["Sheet"]
        
        # Save with preserved structure
        output_wb.save(output_path)
        original_wb.close()
        output_wb.close()
    except Exception:
        # If formatting preservation fails, just use the basic export
        pass
