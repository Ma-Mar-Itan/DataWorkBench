"""
Export engine.

Walks the workbook and applies cleaning rules through the ``RuleIndex``.
All safety invariants that the original translation exporter guaranteed
still hold here, because the rule engine enforces them at a lower layer:

- Whole-cell matching only. The engine never performs substring matching,
  so shorter values never leak into longer phrases.
- Formula cells are skipped. Their formula string is never normalized,
  matched, or written.
- Non-string scalars are untouched. Numbers, dates, booleans, None pass through.
- Blank (``SET_BLANK``) writes ``None`` — a true empty cell, not the string "".
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.normalizer import trim_and_collapse
from core.rules_engine import CellContext, RuleIndex
from models.schemas import CleaningRule


# --------------------------------------------------------------------------- #

@dataclass
class ExportResult:
    """Summary of an export run."""

    output_bytes: bytes
    cells_visited: int
    cells_replaced: int
    cells_blanked: int
    cells_skipped_formula: int
    sheet_count: int
    # rule_id -> how many cells that rule touched
    per_rule_counts: Dict[str, int] = field(default_factory=dict)

    @property
    def total_changed(self) -> int:
        return self.cells_replaced + self.cells_blanked


# --------------------------------------------------------------------------- #

def _is_formula_cell(cell: Cell) -> bool:
    if cell.data_type == "f":
        return True
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


def _headers_for(ws: Worksheet) -> Dict[int, str]:
    """Same header derivation as the extractor — must stay in sync."""
    headers: Dict[int, str] = {}
    if ws.max_row and ws.max_row >= 1:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), ())
        for cell in first_row:
            v = cell.value
            if isinstance(v, str) and v.strip():
                headers[cell.column] = trim_and_collapse(v)
    return headers


# --------------------------------------------------------------------------- #

def apply_rules(file_bytes: bytes, rules: List[CleaningRule]) -> ExportResult:
    """Apply ``rules`` to a workbook and return the cleaned bytes.

    Parameters
    ----------
    file_bytes:
        Raw bytes of an .xlsx file.
    rules:
        List of ``CleaningRule`` objects. Disabled or invalid rules are
        skipped by the ``RuleIndex`` constructor.
    """
    index = RuleIndex(rules)
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=False, read_only=False)

    cells_visited = 0
    cells_replaced = 0
    cells_blanked = 0
    formula_skipped = 0
    per_rule_counts: Dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = _headers_for(ws)

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                cells_visited += 1
                value = cell.value

                # Non-strings never match; skip without cost.
                if not isinstance(value, str):
                    continue

                if _is_formula_cell(cell):
                    formula_skipped += 1
                    continue

                col_label = headers.get(cell.column) or get_column_letter(cell.column)
                ctx = CellContext(sheet=sheet_name, column=col_label, raw_value=value)

                application = index.apply(ctx)
                if application is None:
                    continue

                if application.clear_cell:
                    cell.value = None
                    cells_blanked += 1
                else:
                    cell.value = application.target_value
                    cells_replaced += 1

                per_rule_counts[application.rule_id] = (
                    per_rule_counts.get(application.rule_id, 0) + 1
                )

    buf = BytesIO()
    wb.save(buf)
    sheet_count = len(wb.sheetnames)
    wb.close()

    return ExportResult(
        output_bytes=buf.getvalue(),
        cells_visited=cells_visited,
        cells_replaced=cells_replaced,
        cells_blanked=cells_blanked,
        cells_skipped_formula=formula_skipped,
        sheet_count=sheet_count,
        per_rule_counts=per_rule_counts,
    )


# --------------------------------------------------------------------------- #

def suggest_output_filename(original: Optional[str]) -> str:
    """Return ``<stem>_cleaned.xlsx`` from an original filename."""
    if not original:
        return "workbook_cleaned.xlsx"
    base = original.replace("\\", "/").split("/")[-1]
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    elif "." in base:
        base = base.rsplit(".", 1)[0]
    if not base:
        base = "workbook"
    return f"{base}_cleaned.xlsx"
