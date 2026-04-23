"""
Preview builder.

Renders a sheet's first N rows as two parallel DataFrames — original and
cleaned — plus a "changed" boolean mask so the UI can highlight
transformations. Preview logic is driven by the exact same ``RuleIndex``
the exporter uses.
"""

from __future__ import annotations

from io import BytesIO
from typing import List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.exporter import _headers_for
from core.rules_engine import CellContext, RuleIndex
from models.schemas import CleaningRule


DEFAULT_PREVIEW_ROWS = 100


def list_sheet_names(file_bytes: bytes) -> List[str]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=False, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def build_preview(
    file_bytes: bytes,
    sheet_name: str,
    rules: List[CleaningRule],
    max_rows: int = DEFAULT_PREVIEW_ROWS,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Return (original, cleaned, changed_mask) DataFrames for one sheet.

    The third DataFrame is same-shape booleans — True where the cleaned
    value differs from the original. The UI uses it to colour changed cells.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=False, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in workbook.")
        ws = wb[sheet_name]
        headers = _headers_for(ws)
        index = RuleIndex(rules)

        orig_rows: list[list[object]] = []
        clean_rows: list[list[object]] = []
        changed_rows: list[list[bool]] = []

        for i, row in enumerate(ws.iter_rows(values_only=False)):
            if i >= max_rows:
                break
            orig_row: list[object] = []
            clean_row: list[object] = []
            changed_row: list[bool] = []
            for cell in row:
                val = cell.value
                orig_row.append(val)

                # Non-strings pass through untouched.
                if not isinstance(val, str):
                    clean_row.append(val)
                    changed_row.append(False)
                    continue

                # Formulas pass through untouched.
                if cell.data_type == "f" or val.startswith("="):
                    clean_row.append(val)
                    changed_row.append(False)
                    continue

                col_label = headers.get(cell.column) or get_column_letter(cell.column)
                ctx = CellContext(sheet=sheet_name, column=col_label, raw_value=val)
                application = index.apply(ctx)

                if application is None:
                    clean_row.append(val)
                    changed_row.append(False)
                elif application.clear_cell:
                    clean_row.append(None)
                    changed_row.append(True)
                else:
                    clean_row.append(application.target_value)
                    changed_row.append(application.target_value != val)

            orig_rows.append(orig_row)
            clean_rows.append(clean_row)
            changed_rows.append(changed_row)

        if not orig_rows:
            empty = pd.DataFrame()
            return empty, empty.copy(), empty.copy()

        # Normalize row widths.
        width = max(len(r) for r in orig_rows)
        orig_rows = [r + [None] * (width - len(r)) for r in orig_rows]
        clean_rows = [r + [None] * (width - len(r)) for r in clean_rows]
        changed_rows = [r + [False] * (width - len(r)) for r in changed_rows]

        # Column labels from the header row when possible.
        columns: list[str] = []
        for i in range(width):
            col_idx = i + 1
            columns.append(headers.get(col_idx) or get_column_letter(col_idx))
        # Ensure uniqueness for pandas.
        seen: dict[str, int] = {}
        unique_cols: list[str] = []
        for c in columns:
            if c in seen:
                seen[c] += 1
                unique_cols.append(f"{c} ({seen[c]})")
            else:
                seen[c] = 1
                unique_cols.append(c)

        return (
            pd.DataFrame(orig_rows, columns=unique_cols),
            pd.DataFrame(clean_rows, columns=unique_cols),
            pd.DataFrame(changed_rows, columns=unique_cols),
        )
    finally:
        wb.close()
