"""
Export engine tests — integration with real workbooks.

The Home/Homework test is preserved from v1 and ported onto the new rule
engine. It remains the single clearest proof of the safety invariant.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from core.exporter import apply_rules, suggest_output_filename
from models.schemas import ActionType, CleaningRule, MatchMode, ScopeType
from tests.conftest import AR_HOME, AR_HOMEWORK


# --------------------------------------------------------------------------- #
# The Home / Homework safety proof.
# --------------------------------------------------------------------------- #

def test_arabic_shorter_term_does_not_replace_inside_longer_phrase(
    arabic_home_workbook_bytes: bytes,
) -> None:
    """Map ONLY the shorter Arabic word ('home'). The longer phrase must survive."""
    rules = [
        CleaningRule(
            source_value=AR_HOME,
            target_value="Home",
            action_type=ActionType.REPLACE,
            match_mode=MatchMode.EXACT_RAW,
            scope_type=ScopeType.GLOBAL,
        ),
    ]
    result = apply_rules(arabic_home_workbook_bytes, rules)

    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Sheet1"]

    # Standalone term → translated.
    assert ws.cell(row=2, column=1).value == "Home"
    assert ws.cell(row=4, column=1).value == "Home"

    # Longer phrase → UNCHANGED.
    assert ws.cell(row=3, column=1).value == AR_HOMEWORK

    # Exactly two replacements, not three.
    assert result.cells_replaced == 2


def test_arabic_longer_phrase_can_still_be_mapped_explicitly(
    arabic_home_workbook_bytes: bytes,
) -> None:
    """When the user *does* provide a rule for the longer phrase, it applies."""
    rules = [
        CleaningRule(
            source_value=AR_HOME,
            target_value="Home",
            action_type=ActionType.REPLACE,
            match_mode=MatchMode.EXACT_RAW,
            scope_type=ScopeType.GLOBAL,
        ),
        CleaningRule(
            source_value=AR_HOMEWORK,
            target_value="Homework",
            action_type=ActionType.REPLACE,
            match_mode=MatchMode.EXACT_RAW,
            scope_type=ScopeType.GLOBAL,
        ),
    ]
    result = apply_rules(arabic_home_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Sheet1"]
    assert ws.cell(row=2, column=1).value == "Home"
    assert ws.cell(row=3, column=1).value == "Homework"
    assert ws.cell(row=4, column=1).value == "Home"
    assert result.cells_replaced == 3


# --------------------------------------------------------------------------- #
# Survey workbook — realistic end-to-end scenarios.
# --------------------------------------------------------------------------- #

def test_column_scoped_rules_do_not_corrupt_free_text(survey_workbook_bytes: bytes) -> None:
    """A Gender-column rule 'male → 1' must not touch the Note column."""
    rules = [
        CleaningRule(
            source_value="male",
            target_value="1",
            action_type=ActionType.MAP_CODE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.COLUMN,
            scope_sheet="Survey",
            scope_column="Gender",
        ),
    ]
    result = apply_rules(survey_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Survey"]

    # Gender column was recoded.
    assert ws["A2"].value == "1"
    assert ws["A3"].value == "1"   # " Male " — normalized match

    # Note column is intact.
    assert ws["C2"].value == "male student who left early"


def test_global_yes_no_recoding_affects_only_matching_cells(survey_workbook_bytes: bytes) -> None:
    rules = [
        CleaningRule(
            source_value="yes", target_value="1",
            action_type=ActionType.MAP_CODE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
        ),
        CleaningRule(
            source_value="no", target_value="0",
            action_type=ActionType.MAP_CODE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
        ),
    ]
    result = apply_rules(survey_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Survey"]

    # Response column recoded.
    assert ws["B2"].value == "1"    # "Yes"
    assert ws["B3"].value == "0"    # "no"
    assert ws["B4"].value == "1"    # "YES"
    assert ws["B6"].value == "1"    # "yes"

    # The note "some notes about yes" must NOT become "some notes about 1"
    # because the rule uses whole-cell matching.
    assert ws["C4"].value == "some notes about yes"


def test_set_blank_action_produces_empty_cells(survey_workbook_bytes: bytes) -> None:
    rules = [
        CleaningRule(
            source_value="N/A",
            action_type=ActionType.SET_BLANK,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
        ),
        CleaningRule(
            source_value="-",
            action_type=ActionType.SET_BLANK,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
        ),
    ]
    result = apply_rules(survey_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Survey"]

    # Every N/A cell is now None, across all three columns.
    assert ws["A5"].value is None
    assert ws["B5"].value is None
    assert ws["C3"].value is None   # "N/A" in Note
    assert ws["C5"].value is None   # "-" in Note

    assert result.cells_blanked >= 4


def test_formulas_are_never_translated(survey_workbook_bytes: bytes) -> None:
    """The CONCATENATE formula must not be rewritten by any rule."""
    rules = [
        # A rule that would match the formula string if substring matching
        # existed (CONCATENATE is a common substring).
        CleaningRule(
            source_value="CONCATENATE",
            target_value="REPLACED",
            action_type=ActionType.REPLACE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
        ),
    ]
    result = apply_rules(survey_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Survey"]
    # Formula still in place, unchanged.
    assert ws["A8"].value == '=CONCATENATE("Count: ", 5)'
    assert result.cells_skipped_formula >= 1


def test_disabled_rules_do_nothing(survey_workbook_bytes: bytes) -> None:
    rules = [
        CleaningRule(
            source_value="male", target_value="1",
            action_type=ActionType.MAP_CODE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.GLOBAL,
            enabled=False,
        ),
    ]
    result = apply_rules(survey_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    ws = wb["Survey"]
    assert ws["A2"].value == "male"
    assert result.cells_replaced == 0


def test_per_rule_counts_populated(survey_workbook_bytes: bytes) -> None:
    male_rule = CleaningRule(
        source_value="male", target_value="1",
        action_type=ActionType.MAP_CODE,
        match_mode=MatchMode.EXACT_NORMALIZED,
        scope_type=ScopeType.COLUMN,
        scope_sheet="Survey",
        scope_column="Gender",
    )
    female_rule = CleaningRule(
        source_value="female", target_value="2",
        action_type=ActionType.MAP_CODE,
        match_mode=MatchMode.EXACT_NORMALIZED,
        scope_type=ScopeType.COLUMN,
        scope_sheet="Survey",
        scope_column="Gender",
    )
    result = apply_rules(survey_workbook_bytes, [male_rule, female_rule])
    # Fixture: "male", "Male " → 2 male hits; "female", "female" → 2 female.
    assert result.per_rule_counts.get(male_rule.rule_id) == 2
    assert result.per_rule_counts.get(female_rule.rule_id) == 2


def test_sheet_order_preserved(multi_sheet_workbook_bytes: bytes) -> None:
    result = apply_rules(multi_sheet_workbook_bytes, [])
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    assert wb.sheetnames == ["A", "B"]


def test_sheet_scope_confines_replacement(multi_sheet_workbook_bytes: bytes) -> None:
    rules = [
        CleaningRule(
            source_value="male", target_value="1",
            action_type=ActionType.MAP_CODE,
            match_mode=MatchMode.EXACT_NORMALIZED,
            scope_type=ScopeType.SHEET,
            scope_sheet="A",
        ),
    ]
    result = apply_rules(multi_sheet_workbook_bytes, rules)
    wb = load_workbook(BytesIO(result.output_bytes), data_only=False)
    # Sheet A recoded.
    assert wb["A"]["A2"].value == "1"
    # Sheet B untouched.
    assert wb["B"]["A2"].value == "male"


# --------------------------------------------------------------------------- #
# Filename helper
# --------------------------------------------------------------------------- #

def test_suggest_output_filename_basic() -> None:
    assert suggest_output_filename("report.xlsx") == "report_cleaned.xlsx"


def test_suggest_output_filename_strips_paths() -> None:
    assert suggest_output_filename("/tmp/report.xlsx") == "report_cleaned.xlsx"
    assert suggest_output_filename("C:\\docs\\report.xlsx") == "report_cleaned.xlsx"


def test_suggest_output_filename_fallback() -> None:
    assert suggest_output_filename(None) == "workbook_cleaned.xlsx"
    assert suggest_output_filename("") == "workbook_cleaned.xlsx"
