"""
Pytest configuration and shared fixtures.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook


_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Real Arabic test vocabulary for the safety test.
# --------------------------------------------------------------------------- #

# "home" on its own.
AR_HOME = "منزل"
# A longer phrase that begins with the word for "home". This is the exact
# substring-leak risk the safety invariant must prevent.
AR_HOMEWORK = "منزل العمل"


# --------------------------------------------------------------------------- #
# Fixture workbooks
# --------------------------------------------------------------------------- #

@pytest.fixture
def survey_workbook_bytes() -> bytes:
    """A realistic survey-style workbook.

    Columns: Gender | Response | Note

    The "Note" column is free text and contains strings that *start with*
    values the rules will map. If the engine did substring matching, the
    Note column would be corrupted. The safety test verifies that does not
    happen.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    ws["A1"] = "Gender"
    ws["B1"] = "Response"
    ws["C1"] = "Note"

    ws["A2"] = "male";   ws["B2"] = "Yes"; ws["C2"] = "male student who left early"
    ws["A3"] = "Male ";  ws["B3"] = "no";  ws["C3"] = "N/A"
    ws["A4"] = "female"; ws["B4"] = "YES"; ws["C4"] = "some notes about yes"
    ws["A5"] = "N/A";    ws["B5"] = "N/A"; ws["C5"] = "-"
    ws["A6"] = "female"; ws["B6"] = "yes"; ws["C6"] = "free response"

    # A formula — must never be translated or altered.
    ws["A8"] = "=CONCATENATE(\"Count: \", 5)"
    return workbook_to_bytes(wb)


@pytest.fixture
def arabic_home_workbook_bytes() -> bytes:
    """Specialised fixture for the Home/Homework safety test.

    A1: "Label"   B1: "Value"
    A2: منزل       B2: 1
    A3: منزل العمل B3: 2
    A4: منزل       B4: 3
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Label";    ws["B1"] = "Value"
    ws["A2"] = AR_HOME;    ws["B2"] = 1
    ws["A3"] = AR_HOMEWORK;ws["B3"] = 2
    ws["A4"] = AR_HOME;    ws["B4"] = 3
    return workbook_to_bytes(wb)


@pytest.fixture
def multi_sheet_workbook_bytes() -> bytes:
    """Two-sheet workbook used for sheet/column scope tests."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "A"
    ws1["A1"] = "Gender"
    ws1["A2"] = "male"
    ws1["A3"] = "female"

    ws2 = wb.create_sheet("B")
    ws2["A1"] = "Gender"
    ws2["A2"] = "male"
    ws2["A3"] = "female"
    return workbook_to_bytes(wb)
