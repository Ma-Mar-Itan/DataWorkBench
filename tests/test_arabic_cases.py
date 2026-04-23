"""Arabic-specific tests.

Exercises Arabic values through every stage: scan, rule matching,
apply, export. These cases must pass for the product to be safe for
Arabic-speaking users.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook as openpyxl_load

from core.exporter import export_cleaned_workbook
from core.normalizer import normalize
from core.rules_engine import apply_rules
from core.scanner import scan_workbook
from core.workbook_reader import load_workbook_from_bytes
from models.enums import ActionType, MatchMode, ScopeType
from models.schemas import Rule

from ._helpers import make_test_workbook


def _loaded(sheets):
    return load_workbook_from_bytes(make_test_workbook(sheets), filename="ar.xlsx")


def _rule(**kw) -> Rule:
    defaults = dict(
        rule_id="ar1", source_value="", target_value="",
        action_type=ActionType.REPLACE,
        match_mode=MatchMode.EXACT_RAW,
        scope_type=ScopeType.WORKBOOK,
        scope_sheet="", scope_column="",
        enabled=True, created_at=0,
    )
    defaults.update(kw)
    return Rule(**defaults)


class TestArabicScan:
    def test_scan_recognizes_arabic_values(self):
        wb = _loaded({
            "S": [["gender"], ["ذكر"], ["أنثى"], ["ذكر"], ["ذكر"]]
        })
        result = scan_workbook(wb)
        mapped = {v.raw_value: v.total_count for v in result.values}
        assert mapped["ذكر"] == 3
        assert mapped["أنثى"] == 1

    def test_arabic_tatweel_normalizes_to_same_bucket(self):
        # "ذكر" and "ذكــر" (with tatweel) should normalize identically
        assert normalize("ذكر") == normalize("ذكــر")


class TestArabicMatching:
    def test_short_arabic_does_not_affect_longer_phrase(self):
        """The Arabic parallel of the Home/Homework test."""
        wb = _loaded({
            "S": [["c"], ["ذكر"], ["ذكر أحمد الأمر"], ["الذكر الحكيم"]]
        })
        rules = [_rule(source_value="ذكر", target_value="M")]
        result = apply_rules(wb, rules)
        assert len(result.changes) == 1
        assert result.changes[0].before == "ذكر"

        df = wb.dataframes["S"]
        assert df["c"].tolist() == ["M", "ذكر أحمد الأمر", "الذكر الحكيم"]

    def test_tatweel_variants_match_under_normalized(self):
        # With normalized matching, a rule against "ذكر" should match
        # "ذكـــر" (with tatweel) because tatweel is stripped.
        wb = _loaded({"S": [["c"], ["ذكـــر"]]})
        rules = [_rule(source_value="ذكر",
                       target_value="M",
                       match_mode=MatchMode.EXACT_NORMALIZED)]
        result = apply_rules(wb, rules)
        assert result.changed_cells == 1


class TestArabicExport:
    def test_arabic_roundtrips_through_export(self):
        wb = _loaded({
            "البيانات": [["الاسم", "النوع"],
                         ["أحمد", "ذكر"],
                         ["سارة", "أنثى"]]
        })
        rules = [_rule(source_value="ذكر", target_value="M",
                       scope_type=ScopeType.COLUMN,
                       scope_sheet="البيانات",
                       scope_column="النوع")]
        apply_rules(wb, rules)

        blob = export_cleaned_workbook(wb)
        re_wb = openpyxl_load(BytesIO(blob))

        # Sheet name (Arabic) preserved
        assert "البيانات" in re_wb.sheetnames

        ws = re_wb["البيانات"]
        # Headers preserved
        assert ws.cell(row=1, column=1).value == "الاسم"
        assert ws.cell(row=1, column=2).value == "النوع"
        # Row 2: name unchanged, gender replaced
        assert ws.cell(row=2, column=1).value == "أحمد"
        assert ws.cell(row=2, column=2).value == "M"
        # Row 3: unchanged
        assert ws.cell(row=3, column=2).value == "أنثى"
