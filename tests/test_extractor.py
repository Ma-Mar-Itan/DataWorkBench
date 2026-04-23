"""Tests for workbook scanning."""

from __future__ import annotations

from core.extractor import scan_workbook


def test_scan_finds_all_columns(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    cols_by_sheet = result.columns_by_sheet()
    assert "Survey" in cols_by_sheet
    survey_cols = set(cols_by_sheet["Survey"])
    # The header row gave us named columns.
    assert "Gender" in survey_cols
    assert "Response" in survey_cols
    assert "Note" in survey_cols


def test_scan_deduplicates_whitespace_and_case(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    keys = {v.normalized_value for v in result.values}
    # "male", "Male ", and "MALE" should collapse to one entry.
    assert "male" in keys


def test_scan_tracks_column_context(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    male = next(v for v in result.values if v.normalized_value == "male")
    # "male" appears in the Gender column and also inside the Note free-text cell.
    assert "Gender" in male.columns
    # But "male student who left early" is a SEPARATE value (different normalized form).
    assert "male student who left early" in {v.normalized_value for v in result.values}


def test_scan_counts_frequency(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    male = next(v for v in result.values if v.normalized_value == "male")
    # male, Male, (not MALE in this fixture) → 2 occurrences in Gender column.
    assert male.frequency == 2


def test_scan_skips_formula_cells(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    keys = {v.normalized_value for v in result.values}
    # The formula must never enter the registry.
    assert not any("concatenate" in k.lower() for k in keys)
    assert result.formula_cells_skipped >= 1


def test_scan_identifies_missing_tokens(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    missing = [v for v in result.values if v.likely_missing]
    missing_keys = {v.normalized_value for v in missing}
    assert "n/a" in missing_keys
    assert "-" in missing_keys


def test_scan_flags_header_labels(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    header_keys = {v.normalized_value for v in result.values if v.appears_in_headers}
    assert "gender" in header_keys
    assert "response" in header_keys
    assert "note" in header_keys


def test_scan_column_profiles_top_values(survey_workbook_bytes: bytes) -> None:
    result = scan_workbook(survey_workbook_bytes)
    gender_profile = next(
        p for p in result.column_profiles if p.sheet == "Survey" and p.column == "Gender"
    )
    # Gender column should show "male" as a top value.
    top_keys = {raw.lower() for raw, _ in gender_profile.top_values}
    assert "male" in top_keys
    assert gender_profile.unique_count >= 2  # "male" + "female" + maybe "N/A"


def test_scan_handles_workbook_with_no_strings() -> None:
    from openpyxl import Workbook
    from tests.conftest import workbook_to_bytes

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2.5
    result = scan_workbook(workbook_to_bytes(wb))
    assert result.unique_value_count == 0
    assert result.string_cells_scanned == 0
