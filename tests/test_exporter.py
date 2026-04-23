"""Tests for the exporter — sheet names, formulas, Arabic text."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook as openpyxl_load

from core.exporter import export_cleaned_workbook, export_ruleset, export_stats_report
from core.rules_engine import apply_rules
from core.workbook_reader import load_workbook_from_bytes
from models.enums import ActionType, MatchMode, ScopeType
from models.schemas import Rule

from ._helpers import make_test_workbook


def _loaded(sheets):
    return load_workbook_from_bytes(make_test_workbook(sheets), filename="t.xlsx")


def _rule(**kw) -> Rule:
    defaults = dict(
        rule_id="r1", source_value="", target_value="",
        action_type=ActionType.REPLACE,
        match_mode=MatchMode.EXACT_RAW,
        scope_type=ScopeType.WORKBOOK,
        scope_sheet="", scope_column="",
        enabled=True, created_at=0,
    )
    defaults.update(kw)
    return Rule(**defaults)


class TestCleanedWorkbookExport:
    def test_roundtrip(self):
        wb = _loaded({"S": [["c"], ["x"], ["y"]]})
        rules = [_rule(source_value="x", target_value="Y")]
        apply_rules(wb, rules)

        blob = export_cleaned_workbook(wb)
        # Re-read
        re_wb = openpyxl_load(BytesIO(blob))
        ws = re_wb["S"]
        values = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert values == ["Y", "y"]

    def test_sheet_names_preserved(self):
        wb = _loaded({
            "First":  [["c"], ["a"]],
            "Second": [["c"], ["b"]],
        })
        apply_rules(wb, [])
        blob = export_cleaned_workbook(wb)
        re_wb = openpyxl_load(BytesIO(blob))
        assert re_wb.sheetnames == ["First", "Second"]

    def test_arabic_preserved(self):
        wb = _loaded({"S": [["c"], ["أحمد"], ["ذكر"]]})
        apply_rules(wb, [])
        blob = export_cleaned_workbook(wb)
        re_wb = openpyxl_load(BytesIO(blob))
        values = [row[0].value for row in re_wb["S"].iter_rows(min_row=2)]
        assert values == ["أحمد", "ذكر"]

    def test_formulas_preserved_when_not_targeted(self):
        # Build a workbook with a formula and a plain cell
        wb = _loaded({"S": [["v", "f"], [10, "=A2*2"], [20, "=A3*2"]]})
        # Rule targets the value column but not the formula column
        rules = [_rule(source_value="10", target_value="99",
                       match_mode=MatchMode.EXACT_RAW,
                       scope_type=ScopeType.COLUMN,
                       scope_sheet="S", scope_column="v")]
        apply_rules(wb, rules)

        blob = export_cleaned_workbook(wb)
        re_wb = openpyxl_load(BytesIO(blob))
        ws = re_wb["S"]
        # Value column: "10" changed (as a string, because the cell
        # actually contains the string "10" after pandas roundtrip)
        # Formula column: still formulas
        row2 = [ws.cell(row=2, column=1).value, ws.cell(row=2, column=2).value]
        assert row2[1] == "=A2*2", "formula should be preserved"


class TestRulesetExport:
    def test_roundtrip(self):
        from core.ruleset_store import load_ruleset
        rules = [
            _rule(rule_id="x", source_value="a", target_value="A"),
            _rule(rule_id="y", source_value="b", target_value="B",
                  scope_type=ScopeType.SHEET, scope_sheet="S"),
        ]
        blob = export_ruleset(rules, metadata={"author": "test"})
        loaded_rules, meta = load_ruleset(blob)
        assert meta.get("author") == "test"
        assert [r.rule_id for r in loaded_rules] == ["x", "y"]


class TestStatsReport:
    def test_produces_valid_xlsx(self):
        blob = export_stats_report(
            workbook_summary={
                "sheet_count": 1, "total_rows": 2, "total_cols": 1,
                "total_missing": 0, "total_unique_values": 2,
                "per_sheet": [{"sheet": "S", "rows": 2, "cols": 1, "missing": 0}],
            },
            numeric=[],
            categorical=[],
            missing_rows=[],
        )
        # Must parse as a valid workbook
        wb = openpyxl_load(BytesIO(blob))
        assert "Workbook" in wb.sheetnames
