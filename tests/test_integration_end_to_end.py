"""End-to-end smoke test of the full pipeline via the integration module."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook as openpyxl_load

import integration as adapter

from ._helpers import make_test_workbook


def test_full_pipeline_end_to_end():
    # 1. Build a realistic mixed-content workbook
    data = make_test_workbook({
        "Responses": [
            ["id", "consent", "country", "score"],
            [1, "yes",      "U.S.A.",        8],
            [2, "YES",      "USA",           9],
            [3, "No",       "U.S.A.",        5],
            [4, "N/A",      "USA",           7],
            [5, "Yes",      "United States", 6],
        ],
        "Demographics": [
            ["id", "gender"],
            [1, "ذكر"],
            [2, "أنثى"],
            [3, "ذكر"],
        ],
    })

    # 2. Load
    wb = adapter.load(data, "sample.xlsx")
    assert wb.meta.sheet_count == 2
    assert wb.meta.total_rows == 8

    # 3. Scan
    scan = adapter.scan(wb)
    assert scan["summary"]["sheets"] == 2
    # "yes" and "YES" and "Yes" should all appear, plus Arabic entries
    values = {row["value"] for row in scan["values"]}
    assert "yes" in values
    assert "ذكر" in values

    # 4. Define rules
    rule_dicts = [
        {"rule_id": "r-consent", "source_value": "yes", "target_value": "Yes",
         "action_type": "replace", "match_mode": "exact_normalized",
         "scope_type": "column", "scope_sheet": "Responses",
         "scope_column": "consent", "enabled": True},
        {"rule_id": "r-country", "source_value": "U.S.A.", "target_value": "USA",
         "action_type": "replace", "match_mode": "exact_raw",
         "scope_type": "column", "scope_sheet": "Responses",
         "scope_column": "country", "enabled": True},
        {"rule_id": "r-na", "source_value": "N/A", "target_value": "",
         "action_type": "set_blank", "match_mode": "exact_normalized",
         "scope_type": "workbook", "scope_sheet": "",
         "scope_column": "", "enabled": True},
    ]

    # 5. Preview (non-mutating)
    preview_result = adapter.run_preview(wb, rule_dicts)
    assert preview_result["changed_cells"] > 0

    # Verify original was not mutated by preview
    original_consent = wb.dataframes["Responses"]["consent"].tolist()
    assert "yes" in original_consent or "YES" in original_consent

    # 6. Apply (destructive)
    apply_result = adapter.run_apply(wb, rule_dicts)
    assert apply_result["changed_cells"] > 0

    cleaned_consent = wb.dataframes["Responses"]["consent"].tolist()
    # All consent variants should now be "Yes" or blank
    assert all(v in ("Yes", "No", None) or (isinstance(v, float) and v != v)   # NaN
               for v in cleaned_consent), f"unexpected values: {cleaned_consent}"

    # 7. Export cleaned workbook
    cleaned_bytes = adapter.export_workbook_bytes(wb)
    re_wb = openpyxl_load(BytesIO(cleaned_bytes))
    assert re_wb.sheetnames == ["Responses", "Demographics"]

    # Arabic preserved on export
    dem_ws = re_wb["Demographics"]
    assert dem_ws.cell(row=2, column=2).value == "ذكر"

    # 8. Export stats
    stats_bytes = adapter.export_stats_bytes(wb)
    stats_wb = openpyxl_load(BytesIO(stats_bytes))
    assert "Workbook" in stats_wb.sheetnames

    # 9. Export ruleset
    ruleset_bytes = adapter.export_ruleset_bytes(rule_dicts)
    assert b'"schema_version": 1' in ruleset_bytes
    # Arabic survives JSON export (ensure_ascii=False)
    # (At least one rule has Arabic? No in this suite — but we at least
    # verify valid JSON round-trips cleanly.)
    reloaded = adapter.load_rules_json(ruleset_bytes.decode("utf-8"))
    assert len(reloaded) == 3
