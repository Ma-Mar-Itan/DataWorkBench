"""
Workbook reader.

We keep two parallel representations:

  1. A pandas DataFrame per sheet — convenient for scan/stats/type
     inference. Headers are taken from the first row by default.

  2. The raw openpyxl Workbook — preserved unmodified so the exporter
     can write back through it, keeping formulas and formatting intact
     for cells that no rule touches.

Why both?
  * pandas alone can't preserve formulas or non-value formatting on
    export — writing back via openpyxl requires the live workbook.
  * openpyxl alone is slow for analytics; pandas is ~10x faster for
    column-wise work.
"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from models.schemas import SheetMeta, WorkbookMeta


@dataclass
class LoadedWorkbook:
    """A loaded workbook in both representations."""
    filename:   str
    dataframes: dict[str, pd.DataFrame]   # sheet_name -> DataFrame
    wb:         Workbook                  # live openpyxl workbook (data_only=False)
    # second workbook opened with data_only=True, used so formula cells
    # can be identified without losing them from `wb`
    wb_values:  Workbook
    meta:       WorkbookMeta


def load_workbook_from_bytes(
    data: bytes,
    filename: str = "uploaded.xlsx",
    has_header: bool = True,
) -> LoadedWorkbook:
    """Load an .xlsx from raw bytes."""
    return _load(BytesIO(data), filename, has_header)


def load_workbook_from_path(
    path: str,
    has_header: bool = True,
) -> LoadedWorkbook:
    """Load an .xlsx from a filesystem path."""
    with open(path, "rb") as f:
        return _load(BytesIO(f.read()), path.split("/")[-1], has_header)


def _load(buf: BytesIO, filename: str, has_header: bool) -> LoadedWorkbook:
    # Two parallel loads. We rewind the buffer between them.
    raw_bytes = buf.getvalue()

    # For editing/export (keeps formulas as formula strings)
    wb = load_workbook(BytesIO(raw_bytes), data_only=False, keep_vba=False)
    # For inspection (resolves formulas to their last cached value)
    wb_values = load_workbook(BytesIO(raw_bytes), data_only=True, keep_vba=False)

    dataframes: dict[str, pd.DataFrame] = {}
    sheet_metas: list[SheetMeta] = []
    total_rows = 0
    total_cols = 0

    for sheet_name in wb.sheetnames:
        ws = wb_values[sheet_name]
        # Read all rows once; this is fine for typical workbook sizes
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            df = pd.DataFrame()
            headers: list[str] = []
        elif has_header:
            headers = [_header_safe(h, i) for i, h in enumerate(all_rows[0])]
            data_rows = all_rows[1:]
            df = pd.DataFrame(data_rows, columns=headers)
        else:
            headers = [f"col_{i+1}" for i in range(len(all_rows[0]))]
            df = pd.DataFrame(all_rows, columns=headers)

        dataframes[sheet_name] = df
        sheet_metas.append(SheetMeta(
            name=sheet_name,
            rows=len(df),
            cols=len(df.columns),
            headers=headers,
        ))
        total_rows += len(df)
        total_cols  = max(total_cols, len(df.columns))

    meta = WorkbookMeta(
        filename=filename,
        sheet_count=len(wb.sheetnames),
        total_rows=total_rows,
        total_cols=total_cols,
        sheets=sheet_metas,
    )

    return LoadedWorkbook(
        filename=filename,
        dataframes=dataframes,
        wb=wb,
        wb_values=wb_values,
        meta=meta,
    )


def _header_safe(h, i: int) -> str:
    """Coerce a header value to a usable column name, with fallback."""
    if h is None or (isinstance(h, str) and h.strip() == ""):
        return f"col_{i+1}"
    return str(h)


def is_formula_cell(cell) -> bool:
    """
    True if the given openpyxl cell contains a formula.

    Formulas are stored as strings starting with '='.
    """
    v = cell.value
    return isinstance(v, str) and v.startswith("=")
