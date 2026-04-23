"""
Exporter.

Three outputs:

  1. Cleaned workbook (.xlsx) — written via the openpyxl `wb` that the
     rules engine already mutated in place. Formulas, sheet names, and
     non-targeted cells are preserved byte-for-byte where possible.

  2. Statistics report (.xlsx) — a separate workbook summarizing the
     numeric, categorical, and missingness analyses plus the before/after
     delta.

  3. Ruleset JSON — the exact ruleset that produced the output, so it can
     be reapplied to future files.
"""
from __future__ import annotations

import json
from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl import Workbook

from models.schemas import (
    BeforeAfterSummary, CategoricalStats, NumericStats, Rule,
)

from .workbook_reader import LoadedWorkbook


# --------------------------------------------------------------------- #
# Cleaned workbook
# --------------------------------------------------------------------- #
def export_cleaned_workbook(loaded: LoadedWorkbook) -> bytes:
    """
    Serialize the (already mutated) openpyxl workbook to bytes.

    Call `apply_rules(loaded, rules, mutate_workbook=True)` first, then
    call this. Formulas untouched by any rule are preserved.
    """
    buf = BytesIO()
    loaded.wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------- #
# Stats report
# --------------------------------------------------------------------- #
def export_stats_report(
    workbook_summary: dict,
    numeric: list[NumericStats],
    categorical: list[CategoricalStats],
    missing_rows: list[dict],
    before_after: BeforeAfterSummary | None = None,
) -> bytes:
    """Build a fresh .xlsx containing all statistical outputs."""
    wb = Workbook()
    # Sheet 1: workbook summary
    ws = wb.active
    ws.title = "Workbook"
    ws.append(["Metric", "Value"])
    for k in ("sheet_count", "total_rows", "total_cols",
              "total_missing", "total_unique_values"):
        ws.append([k, workbook_summary.get(k)])
    ws.append([])
    ws.append(["Sheet", "Rows", "Cols", "Missing"])
    for s in workbook_summary.get("per_sheet", []):
        ws.append([s["sheet"], s["rows"], s["cols"], s["missing"]])

    # Sheet 2: numeric
    ws2 = wb.create_sheet("Numeric")
    ws2.append(["Sheet", "Column", "Count", "Missing", "Mean",
                "Median", "Std", "Min", "Q1", "Q3", "Max"])
    for n in numeric:
        ws2.append([n.sheet, n.column, n.count, n.missing,
                    n.mean, n.median, n.std,
                    n.min_, n.q1, n.q3, n.max_])

    # Sheet 3: categorical
    ws3 = wb.create_sheet("Categorical")
    ws3.append(["Sheet", "Column", "Non-missing", "Missing",
                "Unique", "Mode", "Top values"])
    for c in categorical:
        top_str = " · ".join(f"{raw} {pct:.1f}%" for raw, _, pct in c.top_values)
        ws3.append([c.sheet, c.column, c.non_missing, c.missing,
                    c.unique, c.mode, top_str])

    # Sheet 4: missingness
    ws4 = wb.create_sheet("Missingness")
    ws4.append(["Sheet", "Column", "Missing", "Total", "Percent"])
    for m in missing_rows:
        ws4.append([m["sheet"], m["column"], m["missing"], m["total"],
                    round(m["pct"], 2)])

    # Sheet 5: before/after (if provided)
    if before_after is not None:
        ws5 = wb.create_sheet("Before vs After")
        ws5.append(["Metric", "Value"])
        ws5.append(["Changed cells",    before_after.changed_cells])
        ws5.append(["Affected sheets",  before_after.affected_sheets])
        ws5.append(["Affected columns", before_after.affected_columns])
        ws5.append(["Cells blanked",    before_after.missing_added])
        ws5.append([])
        ws5.append(["Column", "Before unique", "After unique"])
        for col, (b, a) in before_after.category_reduction.items():
            ws5.append([col, b, a])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------- #
# Ruleset JSON
# --------------------------------------------------------------------- #
def export_ruleset(
    rules: Iterable[Rule],
    metadata: dict | None = None,
) -> bytes:
    """Serialize rules + optional metadata as pretty-printed JSON bytes."""
    payload = {
        "schema_version": 1,
        "metadata":       metadata or {},
        "rules":          [r.to_dict() for r in rules],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
