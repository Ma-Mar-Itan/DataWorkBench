"""
Workbook scanner: extract unique values with column-level context.

Unlike the original translation scanner — which only cared that a value
contained Arabic — this one captures *every* string cell and records
which sheet and which column (by header label) each value appears in.
That column context is what makes rules like
"in the Gender column, map 'male' to 1" possible.

How header labels are derived
-----------------------------
We treat the **first row of each sheet** as the header row. This matches
how almost every tabular Excel file is laid out. If a sheet has no first
row (empty sheet), columns fall back to spreadsheet letters (A, B, C…).

If a sheet's first row contains empty cells, those columns also fall back
to letters. A cell above a blank header cell still gets a column identity
so it can be scoped correctly.

Formula safety
--------------
Formula cells are skipped entirely. Their formula text is never added to
the value registry, never classified, never eligible for replacement.

Header-row values
-----------------
First-row cells *are* scanned and flagged ``appears_in_headers=True``.
The user can write rules against headers too — scope them with the
``SHEET`` scope and the appropriate match mode — or they can choose to
ignore them in the UI.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional, Set

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.classifier import classify_value, is_likely_missing
from core.normalizer import normalize_value, trim_and_collapse
from models.schemas import ExtractedValue, ValueClass


# Cap on per-value metadata storage — enough for UI context, small enough
# to not blow up memory on large workbooks with chatty free-text columns.
_MAX_SAMPLE_LOCATIONS = 5
_MAX_COLUMNS_TRACKED = 20
_MAX_SHEETS_TRACKED = 20


# --------------------------------------------------------------------------- #
# Column profile — per (sheet, column) stats for the Value Explorer.
# --------------------------------------------------------------------------- #

@dataclass
class ColumnProfile:
    """Per-column statistics from one sheet.

    Populated as we scan. The top-k repeated values are materialised at
    the end so we don't carry a full Counter into the rest of the app.
    """

    sheet: str
    column: str                              # header label, or "A"/"B"… fallback
    cell_count: int = 0
    non_empty_count: int = 0
    string_count: int = 0
    numeric_count: int = 0
    missing_like_count: int = 0
    unique_values: Set[str] = field(default_factory=set)       # normalized
    value_frequency: Dict[str, int] = field(default_factory=dict)   # normalized -> count
    top_values: List[tuple[str, int]] = field(default_factory=list)  # (raw repr, count)

    @property
    def unique_count(self) -> int:
        return len(self.unique_values)

    def finalize(self, raw_by_norm: Dict[str, str], k: int = 5) -> None:
        """Materialise ``top_values`` as the k most common values.

        ``raw_by_norm`` maps normalized → representative raw text so the UI
        can show "Male" rather than "male".
        """
        items = sorted(self.value_frequency.items(), key=lambda kv: (-kv[1], kv[0]))
        self.top_values = [(raw_by_norm.get(n, n), c) for n, c in items[:k]]


# --------------------------------------------------------------------------- #
# ScanResult — top-level scan output.
# --------------------------------------------------------------------------- #

@dataclass
class ScanResult:
    """Result of a workbook scan."""

    values: List[ExtractedValue] = field(default_factory=list)
    column_profiles: List[ColumnProfile] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)
    total_cells_scanned: int = 0
    string_cells_scanned: int = 0
    formula_cells_skipped: int = 0

    @property
    def sheet_count(self) -> int:
        return len(self.sheet_names)

    @property
    def unique_value_count(self) -> int:
        return len(self.values)

    @property
    def missing_token_count(self) -> int:
        return sum(1 for v in self.values if v.likely_missing)

    @property
    def header_label_count(self) -> int:
        return sum(1 for v in self.values if v.appears_in_headers)

    @property
    def likely_categorical_count(self) -> int:
        """Values classified as likely categorical (repeated short text)."""
        return sum(
            1 for v in self.values
            if v.value_class in (ValueClass.TEXT_CATEGORY, ValueClass.NUMERIC_LIKE)
            and v.frequency >= 2
        )

    def columns_by_sheet(self) -> Dict[str, List[str]]:
        """Return {sheet_name: [column_labels...]} in scan order."""
        out: Dict[str, List[str]] = {s: [] for s in self.sheet_names}
        for p in self.column_profiles:
            out.setdefault(p.sheet, [])
            if p.column not in out[p.sheet]:
                out[p.sheet].append(p.column)
        return out


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _is_formula_cell(cell: Cell) -> bool:
    """True if the cell holds a formula."""
    if cell.data_type == "f":
        return True
    val = cell.value
    if isinstance(val, str) and val.startswith("="):
        return True
    return False


def _cell_ref(sheet_name: str, cell: Cell) -> str:
    return f"{sheet_name}!{cell.coordinate}"


def _derive_header_labels(ws: Worksheet) -> Dict[int, str]:
    """Return {column_index (1-based): header_label} for this sheet.

    Uses the first row's cell values where non-empty and string-ish, and
    falls back to spreadsheet letters (A, B, C…) where missing.
    """
    headers: Dict[int, str] = {}
    if ws.max_row and ws.max_row >= 1:
        # iter_rows is more robust than ws[1] for sparse sheets.
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), ())
        for cell in first_row:
            col_idx = cell.column
            val = cell.value
            if isinstance(val, str) and val.strip():
                headers[col_idx] = trim_and_collapse(val)
    # Fallback for any column index we encounter later that has no header.
    return headers


# --------------------------------------------------------------------------- #
# Scan
# --------------------------------------------------------------------------- #

def scan_workbook(file_bytes: bytes) -> ScanResult:
    """Scan an .xlsx's bytes and return a populated ScanResult.

    The scan walks every cell of every sheet (including hidden ones),
    skipping formulas. Each unique normalized string value aggregates into
    one ``ExtractedValue``; per-column stats collect separately into
    ``ColumnProfile`` objects.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=False, read_only=False)

    values: Dict[str, ExtractedValue] = {}          # normalized -> ExtractedValue
    raw_by_norm: Dict[str, str] = {}                # normalized -> representative raw
    profiles: Dict[tuple[str, str], ColumnProfile] = {}

    total_cells = 0
    string_cells = 0
    formula_skipped = 0

    sheet_names = list(wb.sheetnames)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = _derive_header_labels(ws)

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for cell in row:
                total_cells += 1
                col_idx = cell.column
                col_label = headers.get(col_idx) or get_column_letter(col_idx)
                profile_key = (sheet_name, col_label)
                profile = profiles.get(profile_key)
                if profile is None:
                    profile = ColumnProfile(sheet=sheet_name, column=col_label)
                    profiles[profile_key] = profile
                profile.cell_count += 1

                value = cell.value

                # Numeric / bool / datetime / None → update profile but skip value registry.
                if value is None:
                    continue
                profile.non_empty_count += 1
                if isinstance(value, bool) or isinstance(value, (int, float)):
                    profile.numeric_count += 1
                    continue
                if not isinstance(value, str):
                    # Datetime etc. — no string value to classify for recoding.
                    continue

                # Formula safety — never touch formulas.
                if _is_formula_cell(cell):
                    formula_skipped += 1
                    continue

                string_cells += 1
                profile.string_count += 1

                normalized = normalize_value(value)
                if not normalized:
                    # Whitespace-only after normalization — treat as empty.
                    continue

                # Track in per-column profile.
                profile.unique_values.add(normalized)
                profile.value_frequency[normalized] = profile.value_frequency.get(normalized, 0) + 1
                if is_likely_missing(normalized):
                    profile.missing_like_count += 1

                # Track in global value registry.
                is_header_cell = (row_idx == 1)
                existing = values.get(normalized)
                if existing is None:
                    raw_display = trim_and_collapse(value)
                    raw_by_norm[normalized] = raw_display
                    values[normalized] = ExtractedValue(
                        raw_value=raw_display,
                        normalized_value=normalized,
                        frequency=1,
                        sheets=[sheet_name],
                        columns=[col_label],
                        sample_locations=[_cell_ref(sheet_name, cell)],
                        likely_missing=is_likely_missing(normalized),
                        appears_in_headers=is_header_cell,
                    )
                else:
                    existing.frequency += 1
                    if sheet_name not in existing.sheets and len(existing.sheets) < _MAX_SHEETS_TRACKED:
                        existing.sheets.append(sheet_name)
                    if col_label not in existing.columns and len(existing.columns) < _MAX_COLUMNS_TRACKED:
                        existing.columns.append(col_label)
                    if len(existing.sample_locations) < _MAX_SAMPLE_LOCATIONS:
                        existing.sample_locations.append(_cell_ref(sheet_name, cell))
                    if is_header_cell:
                        existing.appears_in_headers = True

    # Classification pass — once per unique value.
    for ev in values.values():
        ev.value_class = classify_value(
            raw_value=ev.raw_value,
            normalized_value=ev.normalized_value,
            frequency=ev.frequency,
            appears_in_headers=ev.appears_in_headers,
        )

    # Finalize column profiles.
    for prof in profiles.values():
        prof.finalize(raw_by_norm)

    wb.close()

    # Sort outputs for deterministic UI rendering:
    # values: high-frequency first, then alphabetical fallback.
    sorted_values = sorted(values.values(), key=lambda v: (-v.frequency, v.normalized_value))
    # profiles: sheet order matches workbook, columns within sheet alphabetised.
    sorted_profiles: List[ColumnProfile] = []
    for sname in sheet_names:
        same_sheet = [p for p in profiles.values() if p.sheet == sname]
        same_sheet.sort(key=lambda p: p.column)
        sorted_profiles.extend(same_sheet)

    return ScanResult(
        values=sorted_values,
        column_profiles=sorted_profiles,
        sheet_names=sheet_names,
        total_cells_scanned=total_cells,
        string_cells_scanned=string_cells,
        formula_cells_skipped=formula_skipped,
    )
